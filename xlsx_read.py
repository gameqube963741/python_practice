import openpyxl

workbook = openpyxl.load_workbook('test.xlsx')

sheet = workbook.worksheets[0]

print(sheet['A1'],sheet['A1'].value)

print(sheet.max_row, sheet.max_column)

for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value,end=' ')
    print()
sheet['A3']='Perry'
workbook.save('test.xlsx')